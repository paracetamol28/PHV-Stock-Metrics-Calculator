!pip install yfinance
!pip install pandas
!pip install openpyxl

import pandas as pd
import datetime
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Font, Alignment
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import yfinance as yf


class StockMetricsCalculator:
    def __init__(self, stocks, metrics):
        self.stocks = stocks
        self.metrics = metrics


    def fetch_historical_data(self):
        data = yf.download(self.stocks, period="1d", interval="1d")
        return data


    def select_latest_data(self, data):
        latest_data = data.iloc[-1].to_frame().transpose()
        return latest_data


    def calculate_metrics(self, latest_data):
      metric_values = {}
      for metric, column in self.metrics.items():
        if column:
            if column in latest_data.columns:
                values = latest_data[column].values.tolist()
            else:
                values = ['N/A'] * len(self.stocks)
        else:
            values = self.calculate_custom_metric(metric)
        metric_values[metric] = values
      return metric_values


    def calculate_custom_metric(self, metric):
      if metric == 'Last close price ($)':
        last_close_price = self.get_last_close_price()
        return last_close_price
      elif metric == 'Operating Margin (%)':
        operating_margin = self.get_operating_margin()
        return operating_margin
      elif metric == 'Revenue - 1 year annualized growth (%)':
        revenue_growth = self.get_revenue_growth()
        return revenue_growth
      elif metric == 'Net Income - 1 year annualized growth (%)':
        net_income_growth = self.get_net_income_growth()
        return net_income_growth
      elif metric == 'Stock YTD performance':
        ytd_return = self.get_ytd_return()
        return ytd_return
      elif metric == 'Short interest (%)':
        short_interest = self.get_short_interest()
        return short_interest
      elif metric == 'EV / (EBITDA - Capex)':
        ev_over_ebitda_minus_capex = self.get_ev_over_ebitda_minus_capex()
        return ev_over_ebitda_minus_capex
      else:
        return ['N/A'] * len(self.stocks)


    def get_operating_margin(self):
        operating_margin_values = []
        for stock in self.stocks:
            data = yf.Ticker(stock)
            info = data.info
            operating_margin = info.get('operatingMargins', 'N/A')
            if operating_margin != 'N/A':
                operating_margin_values.append(operating_margin * 100)
            else:
                operating_margin_values.append(operating_margin)
        return operating_margin_values


    def get_last_close_price(self):
        last_close_price_values = []
        for stock in self.stocks:
            data = yf.Ticker(stock)
            info = data.info
            last_close_price = info.get('previousClose', 'N/A')
            last_close_price_values.append(last_close_price)
        return last_close_price_values


    def get_revenue_growth(self):
        revenue_growth_values = []
        for stock in self.stocks:
            data = yf.Ticker(stock)
            info = data.info
            revenue_growth = info.get('revenueGrowth', 'N/A')
            if revenue_growth != 'N/A':
                revenue_growth_values.append(revenue_growth * 100)
            else:
                revenue_growth_values.append(revenue_growth)
        return revenue_growth_values


    def get_net_income_growth(self):
        net_income_growth_values = []
        for stock in self.stocks:
            data = yf.Ticker(stock)
            info = data.info
            net_income_growth = info.get('earningsGrowth', 'N/A')
            if net_income_growth != 'N/A':
                net_income_growth_values.append(net_income_growth * 100)
            else:
                net_income_growth_values.append(net_income_growth)
        return net_income_growth_values


    def get_ev_over_ebitda_minus_capex(self):
        ev_over_ebitda_minus_capex_values = []
        for stock in self.stocks:
            data = yf.Ticker(stock)
            info = data.info
            capex = info.get('freeCashflow', 'N/A') - info.get('operatingCashflow', 'N/A')
            ev_over_ebitda_minus_capex = info.get('enterpriseValue', 'N/A')/(info.get('ebitda', 'N/A') - capex)
            ev_over_ebitda_minus_capex_values.append(ev_over_ebitda_minus_capex)
        return ev_over_ebitda_minus_capex_values   


    def get_ytd_return(self):
        today = datetime.date.today()
        start_date = datetime.date(today.year, 1, 1)
        ytd_return_values = []
        for stock in self.stocks:
            ticker = yf.Ticker(stock)
            data = ticker.history(start=start_date, end=today)
            ytd_return = (data['Close'][-1] - data['Close'][0]) / data['Close'][0]
            if ytd_return != 'N/A':
              ytd_return_values.append(ytd_return*100)
            else :
              ytd_return_values.append(ytd_return)
        return ytd_return_values


    def get_short_interest(self):
        short_interest_values = []
        for stock in self.stocks:
            data = yf.Ticker(stock)
            info = data.info
            short_interest = info.get('shortPercentOfFloat', 'N/A')
            if short_interest != 'N/A':
              short_interest_values.append(short_interest*100)
            else :
              short_interest_values.append(short_interest)
        return short_interest_values
   

    def create_excel_sheet(self, metric_values, decimal_digits=2):
      workbook = Workbook()
      sheet = workbook.active
      headers = ['Metric'] + self.stocks
      sheet.append(headers)
      header_row = sheet[1]
      header_font = Font(bold=True)
      header_alignment = Alignment(horizontal='center')
      for cell in header_row:
          cell.font = header_font
          cell.alignment = header_alignment
      for metric, values in metric_values.items():
          if metric == 'EV / (EBITDA - Capex)':
              values = [f"x{round(value, decimal_digits)}" if value != 'N/A' else value for value in values]
          elif metric in ['Stock YTD performance', 'Revenue - 1 year annualized growth (%)',
                          'Net Income - 1 year annualized growth (%)', 'Short interest (%)',
                          'Operating Margin (%)']:
              values = [f"{round(value, decimal_digits)}%" if value != 'N/A' else value for value in values]
          else:
              values = [round(value, decimal_digits) if value != 'N/A' else value for value in values]
          row = [metric] + values
          sheet.append(row)

      # Add the date for the calculated metrics
      today = date.today().strftime("%Y-%m-%d")
      date_row = ['Date for the calculated metrics (today)', today]
      sheet.append(date_row)

      for column in sheet.columns:
          max_length = 0
          column_letter = column[0].column_letter
          for cell in column:
              try:
                  if len(str(cell.value)) > max_length:
                      max_length = len(cell.value)
              except:
                  pass
          adjusted_width = (max_length + 2) * 1.2
          sheet.column_dimensions[column_letter].width = adjusted_width

      # Left-align all cells
      for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
          for cell in row:
              cell.alignment = Alignment(horizontal='left')

      workbook.save(filename='stock_metrics.xlsx')


    def print_metrics(self, metric_values):
        for metric, value in metric_values.items():
            print(f"{metric}: {value}")


    def send_email_with_attachment(self, recipient_email, subject, body, attachment_path):
        sender_email = 'your_email@example.com'
        sender_password = 'your_password'

        message = MIMEMultipart()
        message['From'] = sender_email
        message['To'] = recipient_email
        message['Subject'] = subject

        message.attach(MIMEText(body, 'plain'))

        with open(attachment_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={attachment_path}')
            message.attach(part)

        try:
            server = smtplib.SMTP('smtp.gmail.com', 587)
            server.starttls()
            server.login(sender_email, sender_password)
            server.send_message(message)
            server.quit()
            print("Email sent successfully!")
        except Exception as e:
            print("Error sending email:", str(e))


stocks = ['GOOGL', 'AMZN', 'AAPL']
metrics = {
    'S&P 500 weight (%)': '', 
    'Last close price ($)': '',
    'Operating Margin (%)': '',  
    'EV / (EBITDA - Capex)': '',  
    'Stock YTD performance': '',  
    'Revenue - 1 year annualized growth (%)': '',  
    'Net Income - 1 year annualized growth (%)': '',  
    'Short interest (%)': ''
}


calculator = StockMetricsCalculator(stocks, metrics)
data = calculator.fetch_historical_data()
latest_data = calculator.select_latest_data(data)
metric_values = calculator.calculate_metrics(latest_data)

print('stocks: [Google, Amazon, Apple]')
calculator.print_metrics(metric_values)

from google.colab import files
calculator.create_excel_sheet(metric_values, decimal_digits=2)
files.download('stock_metrics.xlsx')

recipient_email = 'recipient_email@example.com'
subject = 'Stock Metrics Report'
body = 'Please find attached the stock metrics report.'
attachment_path = 'stock_metrics.xlsx'
calculator.send_email_with_attachment(recipient_email, subject, body, attachment_path)
